import requests
import openpyxl
import datetime
import os

cookies='s=de1219jmoz; device_id=e73beb23a8ee957bc53f8f889b257a67; __utmz=1.1576089993.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); bid=f1d82844243778c4a52081ddfab3750b_k4wd2ajx; _ga=GA1.2.12831529.1577947531; remember.sig=K4F3faYzmVuqC0iXIERCQf55g2Y; xq_a_token.sig=QNf9qVYVCuGxsgSs6mR0GfTqJ0I; xqat.sig=vm1MpHMX74eIHWOCBtH2XWu7C8M; xq_r_token.sig=fUpHV1vf4q3jalTr-lYMDdaAGRg; xq_is_login.sig=J3LxgPVPUzbBg3Kee_PquUfih7Q; u.sig=D9BmdCGQ_H3H6IjLy6RO1jJEAps; acw_tc=2760822c15791581845323757efd98cecb0eb0346c1ff2f69e781420f1b61d; cookiesu=391579164030751; remember=1; xq_a_token=2bf7e91f15aab4bfd9ad1ad1038d7fe376207e96; xqat=2bf7e91f15aab4bfd9ad1ad1038d7fe376207e96; xq_r_token=0fb4a495b09a298bdfa88ff1a0986e8e89e1e3ed; xq_is_login=1; u=8387456080; aliyungf_tc=AQAAACAbmSnlNA8ANtVo33w6G8Wju2rQ; snbim_minify=true; __utma=1.808383967.1576089993.1579247322.1579483600.24; __utmc=1; __utmt=1; Hm_lvt_1db88642e346389874251b5a1eded6e3=1579189308,1579243134,1579483573,1579484999; Hm_lpvt_1db88642e346389874251b5a1eded6e3=1579485019; __utmb=1.8.10.1579483600'
headers = {
'Cookie':cookies,
'elastic-apm-traceparent': '00-eaa7525216525701aa97ca8da2ede06f-217816b003518ed9-01',
'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.117 Safari/537.36'
}

class etf():
    def __init__(self):
        self.update_session()

    def update_session(self):
        session = requests.Session()
        # 伪造请求头才能登录首页
        session.headers = headers
        # 登录首页，目的是记录连接方式（包括cookies）
        session.get('https://xueqiu.com/')
        self.session = session

    def get_follower_num(self,etf_name):
        url='https://xueqiu.com/recommend/pofriends.json?type=1&code={}&start=0&count=815&_=1579485048158'.format(etf_name)
        session = self.session
        r = session.get(url)
        r = r.json()
        return r['totalcount']

# get # of follower
etf_list=['SH515750','SH512710','SH515650','SH512660','SH512880','SH512760','SH515000','SH501090','SH515050','SZ159966']
values=[]
ETF=etf()

for etf in etf_list:
    totalcount=ETF.get_follower_num(etf)
    values.append(totalcount)

# write into an excel
output_path='/Users/priscilla/Desktop/python/Fuguo/ETF_follower'
data_path=output_path+'/雪球ETF关注量统计.xlsx'


if not os.path.exists(data_path):
    print('自己去建个excel！')
    exit
else:
    data = openpyxl.load_workbook(data_path)

table = data[data.sheetnames[0]]
table = data.active
nrows = table.max_row
ncolumns = table.max_column

# current time
now_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
table.cell(nrows+1,1).value=now_time

# etf follower number
col=2
for value in values:
    table.cell(nrows+1,col).value=value
    col+=1
data.save(data_path)

